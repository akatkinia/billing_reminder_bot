import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import pandas as pd
from sqlite3 import connect, PARSE_COLNAMES
from tabulate import tabulate


def save_csv(DB_PATH):
    conn = connect(DB_PATH, isolation_level=None,
                        detect_types=PARSE_COLNAMES)
    db_df = pd.read_sql_query("SELECT * FROM bills", conn)
    db_df.to_csv('./persistant_data/database.csv', index=False, sep=';', encoding='cp1251')
    conn.close


def save_xlsx(DB_PATH):
    conn = connect(DB_PATH, isolation_level=None, detect_types=PARSE_COLNAMES)
    
    # Загружаем данные из таблицы "bills" в DataFrame
    db_df = pd.read_sql_query("SELECT * FROM bills", conn)
    conn.close()

    # Создаем новую книгу Excel и добавляем рабочий лист (workbook, worksheet)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Устанавливаем заголовки столбцов и применяем стили
    border_thin = Border(left=Side(style='thick'), 
                         right=Side(style='thick'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thick'))
    for col_num, col_name in enumerate(db_df.columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DBE5F1", end_color="DBE5F1", fill_type="solid")
        cell.border = border_thin
    
    # Заполняем данными строки
    for row_num, row_data in enumerate(db_df.values, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Применяем границы ко всей таблице
    max_row = len(db_df) + 1
    max_col = len(db_df.columns)
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border

    # Сохраняем файл
    wb.save('./persistant_data/database.xlsx')


def save_txt(DB_PATH):
    conn = connect(DB_PATH, isolation_level=None, detect_types=PARSE_COLNAMES)
    db_df = pd.read_sql_query("SELECT * FROM bills", conn)
    conn.close()

    # Преобразование DataFrame в строку с помощью tabulate
    table_text = tabulate(db_df, headers='keys', tablefmt='grid', showindex=False)

    with open('./persistant_data/database.txt', 'w', encoding='utf-8') as file:
        file.write(table_text)        
